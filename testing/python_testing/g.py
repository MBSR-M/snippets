from datetime import timezone, timedelta, datetime

import mysql
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pytz
from dateutil import parser
from tzlocal import get_localzone

from snippets.database_config import MySQLConfig
from snippets.database_config.mysql_connector import MySQLConnector

# Global counters and master list
count = 0
exception_list = []

db_args = {
    'host': 'localhost',
    'port': 3306,
    'user': 'admin',
    'password': 'admin',
    'database': 'rtcsyncrep',
    'pool_name': 'rtc-events-pool-01',
    'pool_size': 5
}
db = MySQLConfig(**db_args)
db_client = MySQLConnector(db)
db_client.connect()


def format_sheet(sheet_name):
    sheet = workbook[sheet_name]
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(f"Failed to adjust column width, error '{e}'")
                continue
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
        header_cell = column_cells[0]
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_cell.font = Font(bold=True, color="000000")
        header_cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def utc_to_local(utc_time_str):
    # Check if the input is None
    if utc_time_str is None:
        return None
    # Ensure the input is a string
    if not isinstance(utc_time_str, str):
        raise ValueError(f"Expected a string, got {type(utc_time_str)}")
    # Parse the ISO 8601 format timestamp
    try:
        utc_time = parser.isoparse(utc_time_str)
    except ValueError as e:
        raise ValueError(f"Error parsing timestamp: {e}")
    # Ensure the timestamp is localized to UTC
    utc_time = utc_time.replace(tzinfo=pytz.UTC)
    # Get the local system's time zone
    local_tz = get_localzone()
    # Convert to the local time zone
    local_time = utc_time.astimezone(local_tz)
    # Format the local time as a string
    return local_time.strftime('%Y-%m-%d %H:%M:%S')


def parse_timestamp(timestamp):
    """Convert a Unix timestamp to a naive datetime object in IST."""
    ist = timezone(timedelta(hours=5, minutes=30))  # IST is UTC+5:30
    return datetime.fromtimestamp(int(timestamp), tz=ist).replace(tzinfo=None)


def insert_rtc_event_data(self, cur, day_local, msn, node_id, meter_time, nic_time, rtc_drift, text_code,
                          create_time, update_time, event_time):
    rtc_event_data_query = """
        INSERT INTO rtcsyncrep.rtc_messages (day_local, msn, node_id, meter_time, nic_time, rtc_drift, text_code, create_time, update_time, event_time)
        VALUES (TO_DAYS(%s), %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            day_local = TO_DAYS(%s),
            msn = %s, node_id = %s, meter_time = %s, nic_time = %s, rtc_drift = %s, text_code = %s,
            create_time = %s, update_time = %s, event_time = %s;
    """
    try:
        db_client.execute(rtc_event_data_query, (day_local, msn, node_id, meter_time, nic_time, rtc_drift, text_code,
                                                 create_time, update_time, event_time,  # Values for INSERT
                                                 day_local, msn, node_id, meter_time, nic_time, rtc_drift, text_code,
                                                 # Values for UPDATE
                                                 create_time, update_time, event_time))
    except mysql.connector.Error as err:
        if err.errno == 1526:
            self.logger.warning("Partition error: Table has no partition for the provided value.")
        else:
            self.logger.error(f"Error inserting data, MySQL Error: {err}")
    except Exception as e:
        self.logger.error(f"Error inserting data for MSN {msn}: {e}")


def optimized_count_events(df, date_range):
    """
    Optimized function to count events for each 'msn' and date.
    """
    # Create a copy of the DataFrame to avoid SettingWithCopyWarning
    df = df.copy()

    # Extract date from 'event_time'
    df['event_date'] = df['event_time'].dt.date

    # Filter the DataFrame for the relevant date range
    start_date = date_range.min().date()
    end_date = date_range.max().date()
    filtered_df = df[(df['event_date'] >= start_date) & (df['event_date'] <= end_date)]

    # Group by 'msn' and 'event_date' and count occurrences
    grouped = filtered_df.groupby(['msn', 'event_date']).size().reset_index(name='count')

    # Pivot the table to have dates as columns
    pivot_df = grouped.pivot(index='msn', columns='event_date', values='count').fillna(0)

    # Reindex columns to match the full date_range
    pivot_df = pivot_df.reindex(columns=[d.date() for d in date_range], fill_value=0)

    # Reset the index to flatten the DataFrame
    return pivot_df.reset_index()


# Read the Excel file (ensure that headers are present in the file)
df = pd.read_excel(r"C:\Users\MustafaMubashir\Downloads\rtc_data.xlsx")

# Process rows and populate the master list
for index, row in df.iterrows():
    try:
        # Access column 4 using its integer index (4 in this case)
        params_str = row[4]  # Reference column 4 correctly
        if pd.isnull(params_str):
            raise ValueError("Column '4' is null.")

        # Split the string in column '4' by commas and convert it into a list
        parameters = params_str.split(',')
        if len(parameters) < 3:
            raise ValueError("Not enough parameters in column '4'.")
        event_time = pd.to_datetime(row[7])
        data = {
            'day_local': row[1],  # Use integer index 1 for "day_local"
            'msn': row[2],  # Use integer index 2 for "msn"
            'node_id': row[3],  # Use integer index 3 for "node_id"
            'meter_time': parse_timestamp(parameters[1]),
            'nic_time': parse_timestamp(parameters[0]),
            'rtc_drift': int(parameters[2]),
            'text_code': row[5],  # Use integer index 5 for "text_code"
            'create_time': row[6],  # Use integer index 6 for "create_time"
            'update_time': row[7],  # Use integer index 7 for "update_time"
            'event_time': row[8],  # Ensure event_time is in datetime format
        }
        insert_rtc_event_data(cursor, day_local=data['day_local'], msn=data['msn'],
                              node_id=data['node_id'], meter_time=data['meter_time'],
                              nic_time=data['nic_time'],
                              rtc_drift=data['rtc_drift'], text_code=data['text_code'],
                              create_time=utc_to_local(data['create_time']),
                              update_time=utc_to_local(data['update_time']),
                              event_time=utc_to_local(data['event_time']))
    except Exception as e:
        exception_data = {
            'day_local': row[1],  # Use integer index 1 for "day_local"
            'msn': row[2],  # Use integer index 2 for "msn"
            'node_id': row[3],  # Use integer index 3 for "node_id"
            'meter_time': None,
            'nic_time': None,
            'rtc_drift': None,
            'text_code': row[5],  # Use integer index 5 for "text_code"
            'create_time': row[6],  # Use integer index 6 for "create_time"
            'update_time': row[7],  # Use integer index 7 for "update_time"
            'event_time': row[8],  # Ensure event_time is in datetime format
        }
        exception_list.append(exception_data)

print(f"len of master_list : {len(exception_list)}")  # Print the total number of valid rows processed

# Convert the master list to a DataFrame
df_1 = pd.DataFrame(master_list)

# Ensure 'event_time' is converted to datetime type if it's not already
df_1['event_time'] = pd.to_datetime(df_1['event_time'], errors='coerce')

# Define the date range
date_range = pd.date_range("2025-01-01", "2025-02-03", freq="D")

# Filter the DataFrame for relevant text codes
df_sync = df_1[df_1['text_code'] == 'device.RTCSync']
df_set = df_1[df_1['text_code'] == 'device.RTCSet']

# Count events for both sync and set using the optimized function
sync_counts = optimized_count_events(df_sync, date_range)
set_counts = optimized_count_events(df_set, date_range)

# Write to Excel with two sheets

output_path = r"C:\Users\MustafaMubashir\Downloads\output-" + datetime.now().strftime("%Y-%m-%d") + ".xlsx"

with pd.ExcelWriter(output_path) as writer:
    sync_counts.to_excel(writer, sheet_name='RTCSync', index=False)
    set_counts.to_excel(writer, sheet_name='RTCSet', index=False)

workbook = load_workbook(output_path)
for sheet in workbook.sheetnames:
    format_sheet(sheet)
workbook.save(output_path)

print("Excel file created successfully.")
