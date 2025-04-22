import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Constants for paths and date range
FILE_PATH = r"C:\Users\MustafaMubashir\OneDrive - Cyan Connode Ltd\Documents\output11.csv"
OUTPUT_DIR = r"C:\Users\MustafaMubashir\OneDrive - Cyan Connode Ltd\Documents\lastgaspjan2025"
START_DAY = 739617
START_DATE = datetime(2025, 1, 1)
END_DATE = datetime(2025, 1, 31)


file_size_mb = os.path.getsize(FILE_PATH) / (1024 * 1024)
print(f"File size: {file_size_mb:.2f} MB")
column_names = ['id', 'day_local', 'server_time', 'alarm_time', 'msn', 'node_id', 'alarm_status']
df = pd.read_csv(FILE_PATH, header=None, names=column_names)
date_mapping = {
    day: (START_DATE + timedelta(days=day - START_DAY)).strftime('%Y-%m-%d')
    for day in range(START_DAY, START_DAY + 31)
}
df['day_local'] = df['day_local'].map(date_mapping)
df[['server_time', 'alarm_time']] = df[['server_time', 'alarm_time']].apply(
    lambda col: pd.to_datetime(col, errors='coerce', utc=True)
)
df.dropna(subset=['server_time', 'alarm_time'], inplace=True)
df[['server_time', 'alarm_time']] = df[['server_time', 'alarm_time']].apply(
    lambda col: col.dt.tz_convert('Asia/Kolkata')
)
df[['server_time', 'alarm_time']] = df[['server_time', 'alarm_time']].apply(
    lambda col: col.dt.tz_localize(None)
)
filtered_data = df[df['alarm_status'] == 'Last Gasp']
final_columns = ['msn', 'node_id', 'alarm_time', 'server_time', 'alarm_status']
for current_date in pd.date_range(START_DATE, END_DATE):
    date_str = current_date.strftime('%Y-%m-%d')
    formatted_file_name = f"Last-GASP-Alarm-Report-{date_str}.xlsx"
    day_data = filtered_data[filtered_data['day_local'] == date_str]
    day_data = day_data[final_columns]
    day_data.rename(
        columns={'msn': 'MSN', 'node_id': 'Node Id', 'alarm_time': 'Alarm Time', 'server_time': 'Server Time',
                 'alarm_status': 'Alarm Status'}, inplace=True)
    if not day_data.empty:
        output_path = os.path.join(OUTPUT_DIR, formatted_file_name)
        day_data.to_excel(output_path, sheet_name="Last Gasp Data", index=False)
        print(f"File saved: {formatted_file_name}")
        workbook = load_workbook(output_path)
        sheet = workbook['Last Gasp Data']
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(f"Failed to adjust column width, error '{e}'")
                    continue
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width
            header_cell = column_cells[0]
            header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_cell.font = Font(bold=True, color="000000")
            header_cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        workbook.save(output_path)
    else:
        print(f"No data for {date_str}, skipping file creation.")
