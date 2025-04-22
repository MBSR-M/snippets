import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "SI Road Map"

# Define headers and their styles
headers = [
    "Phase",
    "Task",
    "Owner",
    "Start Date",
    "End Date",
    "Status",
    "Jira Ticket",
    "Dependencies",
    "Resources Needed",
    "Comments"
]

header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
header_font = Font(bold=True, color="000000")
alignment = Alignment(horizontal="center", vertical="center")

# Add headers to the sheet
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}1"] = header
    ws[f"{col_letter}1"].fill = header_fill
    ws[f"{col_letter}1"].font = header_font
    ws[f"{col_letter}1"].alignment = alignment

# Add dropdown options for the "Status" column
status_options = ["Not Started", "In Progress", "Completed", "Blocked"]
dv_status = DataValidation(
    type="list",
    formula1=f"{','.join(status_options)}",
    allow_blank=True
)
ws.add_data_validation(dv_status)

# Apply the dropdown to the "Status" column (column F)
for row in range(2, 100):
    dv_status.add(ws[f"F{row}"])

# Add example rows with formulas and logic
example_rows = [
    ["Initiation", "Define Requirements", "John Doe", "2025-01-15", "2025-01-30", "Not Started", "SI-123", "None", "2 Developers, 1 QA", ""],
    ["Planning", "Create Road Map", "Jane Smith", "2025-01-20", "2025-02-05", "In Progress", "SI-124", "Task 1", "Project Manager", "Ensure stakeholder approval"],
    ["Execution", "Develop Features", "Alex Brown", "2025-02-01", "2025-03-01", "Not Started", "SI-125", "Task 2", "Development Team", "Focus on core features"]
]

for row_data in example_rows:
    ws.append(row_data)

# Adjust column widths
for col_num, _ in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = 20

# Save the Excel sheet
output_file = "/mnt/data/SI_Road_Map.xlsx"
wb.save(output_file)
